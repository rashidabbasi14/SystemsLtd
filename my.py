from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import os

sign_on_prefix = True
USER_PREFIX = "S."
role_application = {}
browser_preferences = {}
p_items = []
s_items = []

def main():
    # User Data
    wb = load_workbook(filename = 'UserSheet.xlsx')
    ws = wb.active
    
    # User DMT Sheet
    nwb = Workbook()
    nws = nwb.active
    
    # Browser Preference DMT Sheet
    bwb = Workbook()
    bws = bwb.active
    
    mincol = 1
    maxcol = 1
    minrow = 2
    maxrow = ws.max_row

    init_headers(nws, bws)
    init_error_files()
    init_roles()

    for col in ws.iter_cols(min_row = minrow, min_col = mincol, max_col = maxcol, max_row = maxrow):
        for cell in col:
            if check_errors(ws, cell):
                continue
            
            cell.number_format = '@'
            init_companies(ws,cell)
            nws['A'+str(cell.row)] = ""
            nws['B'+str(cell.row)] = USER_PREFIX + str(ws['B'+str(cell.row)].value).rjust(5,"0")
            nws['C'+str(cell.row)] = str(ws['A'+str(cell.row)].value).upper().strip()
            nws['D'+str(cell.row)] = str(ws['B'+str(cell.row)].value).rjust(5,"0") if not sign_on_prefix else USER_PREFIX.replace(".","") + str(ws['B'+str(cell.row)].value).rjust(5,"0")
            nws['E'+str(cell.row)] = "INT"
            nws['F'+str(cell.row)] = "1"
            nws['G'+str(cell.row)] = set_company_codes(ws,cell)
            nws['H'+str(cell.row)] = ws['D'+str(cell.row)].value
            nws['I'+str(cell.row)] = "20240601M0101"
            nws['J'+str(cell.row)] = datetime.today().strftime('%Y%m%d')
            nws['K'+str(cell.row)] = "20990321"
            nws['L'+str(cell.row)] = "00:00"
            nws['M'+str(cell.row)] = "24:00"
            nws['N'+str(cell.row)] = "5"
            nws['O'+str(cell.row)] = "9"
            nws['P'+str(cell.row)] = set_company_restr(ws,cell)
            nws['Q'+str(cell.row)] = set_application(ws,cell)
            nws['R'+str(cell.row)] = set_function(ws,cell)
            nws['S'+str(cell.row)] = "Y"
            nws['T'+str(cell.row)] = "Y"
            nws['U'+str(cell.row)] = "Y"
            nws['V'+str(cell.row)] = "Y"
            nws['W'+str(cell.row)] = "DDMM"
            nws['X'+str(cell.row)] = "Y"
            nws['Y'+str(cell.row)] = set_override_class(ws,cell)
            nws['Z'+str(cell.row)] = "00"
            nws['AA'+str(cell.row)] = "?."
            nws['AB'+str(cell.row)] = "1"
            
            # BROWSER.PREFERENCES----------------------------
            bws['A'+str(cell.row)] = ""
            bws['B'+str(cell.row)] = USER_PREFIX + str(ws['B'+str(cell.row)].value).rjust(5,"0")
            bws['C'+str(cell.row)] = ""
            bws['D'+str(cell.row)] = "LOCAL"
            bws['E'+str(cell.row)] = browser_preferences[ws['H'+str(cell.row)].value] if ws['H'+str(cell.row)].value in browser_preferences else "ASA.RBHP.BUS.SCREEN"
            # bws['F'+str(cell.row)] = ws['H'+str(cell.row)].value
            
    nwb.save('User DMT.xlsx')
    bwb.save('BP DMT.xlsx')
    save_error_files()
       
def init_headers(nws, bws):
    nws['A1'] = "UPLOAD.COMPANY"
    nws['B1'] = "@ID"
    nws['C1'] = "USER.NAME"
    nws['D1'] = "SIGN.ON.NAME"
    nws['E1'] = "CLASSIFICATION"
    nws['F1'] = "LANGUAGE"
    nws['G1'] = "COMPANY.CODE"
    nws['H1'] = "DEPARTMENT.CODE"
    nws['I1'] = "PASSWORD.VALIDITY"
    nws['J1'] = "START.DATE.PROFILE"
    nws['K1'] = "END.DATE.PROFILE"
    nws['L1'] = "START.TIME"
    nws['M1'] = "END.TIME"
    nws['N1'] = "TIME.OUT.MINUTES"
    nws['O1'] = "ATTEMPTS"
    nws['P1'] = "COMPANY.RESTR"
    nws['Q1'] = "APPLICATION"
    nws['R1'] = "FUNCTION"
    nws['S1'] = "SIGN.ON.OFF.LOG"
    nws['T1'] = "SECURITY.MGMT.L"
    nws['U1'] = "APPLICATION.LOG"
    nws['V1'] = "FUNCTION.ID.LOG"
    nws['W1'] = "INPUT.DAY.MONTH"
    nws['X1'] = "CLEAR.SCREEN"
    nws['Y1'] = "OVERRIDE.CLASS"
    nws['Z1'] = "DEALER.DESK"
    nws['AA1'] = "AMOUNT.FORMAT"
    nws['AB1'] = "DATE.FORMAT"
    
    bws['A1'] = "UPLOAD.COMPANY"
    bws['B1'] = "@ID"
    bws['C1'] = "SKIN.NAME"
    bws['D1'] = "PRINT.LOCATION"
    bws['E1'] = "MAIN.SCREEN"
        
def init_error_files():
    if not os.path.exists("Errors"):  
        os.makedirs("Errors") 
    

def init_roles():
    global role_application
    global browser_preferences
    appl = "ALL.PG" #Other than checker
    c_appl = "@ASA.PK.A.CAD" #Checker
    m_func = "B C D E F H I L P R S V" #Maker
    c_func = "" #Checker
    v_func = "H L P S V" #View Only
    a_func = "A 2 B C D E F H I L P R S V" #All Rights
    
    role_application =	{
        "TO": "ASA.PK.I.TO",
        "GBO": "ASA.PK.I.GBO",
        "OM": "ASA.PK.ALL.OM",
        "BM": "ASA.PK.ALL.BM",
        "AOM": "ASA.PK.A.AOM",
        "ROM": "ASA.PK.A.ROM",
        "HBBO": "ASA.PK.A.HBBO",
        "CAOP": "ASA.PK.I.CAOP",
        "CAOS": "ASA.PK.A.CAOS",
        "CAOM": "ASA.PK.A.CAOM",
        "CTDP": "ASA.PK.I.CTDP",
        "CTDS": "ASA.PK.A.CTDS",
        "CTDM": "ASA.PK.A.CTDM",
        "HCO": "ASA.PK.A.HCO",
        "OSPM": "ASA.PK.V.OSPM",
        "HSPP": "ASA.PK.V.HSPP",
        "HOPS": "ASA.PK.A.HOPS",
        "CCCS": "ASA.PK.I.CCCP",
        "CCCP": "ASA.PK.A.CCCS",
        "CCCM": "ASA.PK.A.CCCM",
        "LO": "ASA.PK.I.LO",
        "ASCM": "ASA.PK.I.ASCM",
        "ARM": "ASA.PK.I.ARM",
        "SCM": "ASA.PK.ALL.SCM",
        "RM": "ASA.PK.ALL.RM",
        "BM": "ASA.PK.A.BM",
        "AM": "ASA.PK.V.AM",
        "RH": "ASA.PK.V.RH",
        "CA": "ASA.PK.A.CA",
        "BA": "ASA.PK.A.BA",
        "ZH": "ASA.PK.V.ZH",
        "MD": "ASA.PK.V.MD",
        "HOD": "ASA.PK.V.HOD",
        "HOB": "ASA.PK.V.HOB",
        "DCEO": "ASA.PK.V.DCEO",
        "CEO": "ASA.PK.V.CEO",
        "IT":  "ASA.PK.ALL.IT"
    }
    
    browser_preferences = {
        "TO": "ASA.RBHP.OPR.TO.SCREEN",
        "GBO": "ASA.RBHP.OPR.GPO.SCREEN",
        "OM": "ASA.RBHP.OPR.OM.SCREEN",
        "BM": "ASA.RBHP.OPR.BM.SCREEN",
        "AOM": "ASA.RBHP.OPR.AOM.SCREEN",
        "ROM": "ASA.RBHP.OPR.ROM.SCREEN",
        "HBBO": "ASA.RBHP.OPR.HBBO.SCREEN",
        "CAOP": "ASA.RBHP.OPR.CAOP.SCREEN",
        "CAOS": "ASA.RBHP.OPR.CAOS.SCREEN",
        "CAOM": "ASA.RBHP.OPR.CAOM.SCREEN",
        "CTDP": "ASA.RBHP.OPR.CTDP.SCREEN",
        "CTDS": "ASA.RBHP.OPR.CTDS.SCREEN",
        "CTDM": "ASA.RBHP.OPR.CTDM.SCREEN",
        "HCO": "ASA.RBHP.OPR.HCO.SCREEN",
        "OSPM": "ASA.RBHP.OPR.OSPM.SCREEN",
        "HSPP": "ASA.RBHP.OPR.HSPP.SCREEN",
        "HOPS": "ASA.RBHP.OPR.HOPS.SCREEN",
        "CCCS": "ASA.RBHP.OPR.CCCS.SCREEN",
        "CCCP": "ASA.RBHP.OPR.CCCP.SCREEN",
        "CCCM": "ASA.RBHP.OPR.CCCM.SCREEN",
        "LO": "ASA.RBHP.BUS.LO.SCREEN",
        "ASCM": "ASA.RBHP.BUS.ASCM.ARM.SCREEN",
        "ARM": "ASA.RBHP.BUS.ASCM.ARM.SCREEN",
        "SCM": "ASA.RBHP.BUS.SCM.SCREEN",
        "RM": "ASA.RBHP.BUS.RM.SCREEN",
        "BM": "ASA.RBHP.BUS.BM.SCREEN",
        "AM": "ASA.RBHP.BUS.AM.SCREEN",
        "RH": "ASA.RBHP.BUS.RH.SCREEN",
        "CA": "ASA.RBHP.BUS.CA.SCREEN",
        "BA": "ASA.RBHP.BUS.BA.SCREEN",
        "ZH": "ASA.RBHP.BUS.HEADS.SCREEN",
        "MD": "ASA.RBHP.BUS.HEADS.SCREEN",
        "HOD": "ASA.RBHP.BUS.HEADS.SCREEN",
        "HOB": "ASA.RBHP.BUS.HEADS.SCREEN",
        "DCEO": "ASA.RBHP.BUS.HEADS.SCREEN",
        "CEO": "ASA.RBHP.BUS.HEADS.SCREEN"
    }
    
def save_error_files():
    Error.f1.close()
    Error.f2.close()
    Error.f3.close()
    Error.f4.close()
    Error.f5.close()
    files = ["Errors/nameError.txt","Errors/signOnNameError.txt","Errors/DAO Warning.txt","Errors/CompanyNameEmpty.txt","Errors/MultipleRoleError.txt"]
    for file in files:
        file_stats = os.stat(file)
        if(file_stats.st_size == 0):
            os.remove(file)
        
        
def check_errors(ws, cell):
    if ws['A'+str(cell.row)].value == None:
        Error.f1.write("ERROR: Row-"+str(cell.row) + ": Name is Empty\n")
    if len(str(ws['B'+str(cell.row)].value)) < 5:
        Error.f2.write("ERROR: Row-"+str(cell.row) + ": Sign On Name String is less than 5\n")
    if len(str(ws['D'+str(cell.row)].value)) > 2:
        Error.f3.write("WARNING Row-"+str(cell.row) + ": DAO Warning (Greater than 2)\n")
    if ws['E'+str(cell.row)].value == None or len(str(ws['E'+str(cell.row)].value)) != 9:
        Error.f4.write("ERROR: Row-"+str(cell.row) + ": Primary Company is Empty or not length of 9\n")
    if ws['C'+str(cell.row)].value == None:
        Error.f5.write("ERROR: Row-"+str(cell.row) + ": Role is Field\n")
        
    return ws['B'+str(cell.row)].value == None
    
    
def set_override_class(ws,cell):
    if ws['H'+str(cell.row)].value in ["TO","GBO"]:
        return "TOGB"
    elif ws['H'+str(cell.row)].value in ["OM","BM","AOM","ROM","HBBO"]:
        return ws['H'+str(cell.row)].value

def init_companies(ws, cell):
    global p_items
    global s_items
    p_items = []
    s_items = []
    
    P_COMPANY = ws['E'+str(cell.row)].value.replace(" ","")
    P_COMPANY = P_COMPANY.replace(",,","")
    P_COMPANY = P_COMPANY.replace(",","::")
    p_items = P_COMPANY.split("::")
    p_items = [x for n, x in enumerate(p_items) if x not in p_items[:n]]
    p_items = [x for n, x in enumerate(p_items) if len(x) == 9 or x == "ALL"]
    
    if ws['F'+str(cell.row)].value != None:
        COMPANY = ws['F'+str(cell.row)].value.replace(" ","")
        COMPANY = COMPANY.replace(",,",",")
        COMPANY = COMPANY.replace(",","::")
        s_items = COMPANY.split("::")
        s_items = [x for n, x in enumerate(s_items) if x not in s_items[:n]]
        s_items = [x for n, x in enumerate(s_items) if x not in p_items]
        s_items = [x for n, x in enumerate(s_items) if len(x) == 9 or x == "ALL"]
        
def set_company_codes(ws,cell):
    global p_items
    global s_items
    return "::".join(p_items + s_items)

def set_company_restr(ws,cell):
    global p_items
    global s_items
    
    if "ALL" not in p_items and "ALL" not in s_items:
        return "::".join(p_items + s_items) + "::ALL"
    else:
        return "::".join(p_items + s_items)
        
def set_application(ws,cell):
    application = []
    for x in p_items:
        application.append("@"+role_application[ws['H'+str(cell.row)].value] if ws['H'+str(cell.row)].value in role_application else "ALL.PG")
        
    for x in s_items:
        application.append("ALL.PG")
        
    if "ALL" not in p_items and "ALL" not in s_items:
        return "::".join(application) + "::ALL.PG"
    else:
        return "::".join(application)   
    
        
def set_function(ws,cell):
    function = []
    for x in p_items:
        function.append("" if ws['H'+str(cell.row)].value in role_application else "H L P S V")
        
    for x in s_items:
        function.append("H L P S V")
    
    if "ALL" not in p_items and "ALL" not in s_items:
        return "::".join(function) + "::H L P S V"
    else:
        return "::".join(function)
    

class Error:
    f1 = open("Errors/nameError.txt", "w")
    f2 = open("Errors/signOnNameError.txt", "w")
    f3 = open("Errors/DAO Warning.txt", "w")
    f4 = open("Errors/CompanyNameEmpty.txt", "w")
    f5 = open("Errors/MultipleRoleError.txt", "w")
    
if __name__ == "__main__":
    main()

